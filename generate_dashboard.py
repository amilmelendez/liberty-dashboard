#!/usr/bin/env python3
"""
Liberty Communications Dashboard Generator
Reads Excel data and generates HTML dashboard with embedded data
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def read_excel_data(excel_file):
    """Read all data from Excel file"""
    try:
        wb = load_workbook(excel_file, data_only=True)
    except FileNotFoundError:
        print(f"Error: Excel file not found: {excel_file}")
        sys.exit(1)
    
    data = {}
    
    # ── PNM MONTHLY ──
    if 'PNM_Monthly' in wb.sheetnames:
        ws = wb['PNM_Monthly']
        labels, values = [], []
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if row[0] is None:
                break
            if row[0] and row[1] is not None:
                labels.append(str(row[0]))
                values.append(float(row[1]) if isinstance(row[1], (int, float)) else 0)
        data['monLabels'] = labels
        data['monVals'] = values
        data['pnmYtdTarget'] = str(ws['C2'].value or '1500')
    
    # ── PNM WEEKLY ──
    if 'PNM_Weekly' in wb.sheetnames:
        ws = wb['PNM_Weekly']
        labels, values = [], []
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if row[0] is None:
                break
            if row[0] and row[1] is not None:
                labels.append(str(row[0]))
                values.append(float(row[1]) if isinstance(row[1], (int, float)) else 0)
        data['wkLabels'] = labels
        data['wkVals'] = values
    
    # ── SOLI5 TAPS ──
    if 'Taps_Weekly' in wb.sheetnames:
        ws = wb['Taps_Weekly']
        labels, values = [], []
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if row[0] is None:
                break
            if row[0] and row[1] is not None:
                labels.append(str(row[0]))
                values.append(float(row[1]) if isinstance(row[1], (int, float)) else 0)
        data['tapsLabels'] = labels
        data['tapsVals'] = values
        data['tapsYtdTarget'] = str(ws['C2'].value or '7000')
    
    # ── REACTIVE LEAKS ──
    if 'Leaks_Weekly' in wb.sheetnames:
        ws = wb['Leaks_Weekly']
        labels, detected, repaired, miles = [], [], [], []
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if row[0] is None:
                break
            if row[0] and row[1] is not None:
                labels.append(str(row[0]))
                detected.append(float(row[1]) if isinstance(row[1], (int, float)) else 0)
                repaired.append(float(row[2]) if isinstance(row[2], (int, float)) else 0)
                miles.append(float(row[3]) if isinstance(row[3], (int, float)) else 0)
        data['lkLabels'] = labels
        data['lkDetected'] = detected
        data['lkRepaired'] = repaired
        data['lkMiles'] = miles
        data['lkYtdTargetDetected'] = str(ws['E2'].value or '')
        data['lkYtdTargetCleared'] = str(ws['E3'].value or '')
    
    # ── REACTIVE MRS ──
    if 'MRs_Weekly' in wb.sheetnames:
        ws = wb['MRs_Weekly']
        labels, completed, cancelled, created = [], [], [], []
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if row[0] is None:
                break
            if row[0] and row[1] is not None:
                labels.append(str(row[0]))
                completed.append(float(row[1]) if isinstance(row[1], (int, float)) else 0)
                cancelled.append(float(row[2]) if isinstance(row[2], (int, float)) else 0)
                created.append(float(row[3]) if isinstance(row[3], (int, float)) else 0)
        data['mrsLabels'] = labels
        data['mrsCompleted'] = completed
        data['mrsCancelled'] = cancelled
        data['mrsCreated'] = created
        data['mrsYtdTarget'] = str(ws['E2'].value or '')
    
    # ── QOE ──
    if 'QOE_Monthly' in wb.sheetnames:
        ws = wb['QOE_Monthly']
        labels, target, actual = [], [], []
        for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
            if row[0] is None:
                break
            if row[0] and row[1] is not None:
                labels.append(str(row[0]))
                target.append(float(row[1]) if isinstance(row[1], (int, float)) else 0)
                actual_val = row[2] if isinstance(row[2], (int, float)) and row[2] > 0 else None
                actual.append(float(actual_val) if actual_val else None)
        data['qoeLabels'] = labels
        data['qoeTarget'] = target
        data['qoeActual'] = actual
    
    # ── SETTINGS ──
    if 'Settings' in wb.sheetnames:
        ws = wb['Settings']
        data['weekLabel'] = str(ws['B1'].value or 'W19 · 2026')
        data['wkNote'] = str(ws['B2'].value or 'W5–W19 = 2026')
        data['pnmTitle'] = str(ws['B3'].value or 'PNM Nodes Completed')
        data['tapsTitle'] = str(ws['B4'].value or 'SOLI5 Taps Cleaned')
        data['leaksTitle'] = str(ws['B5'].value or 'Reactive Leaks Reports')
        data['mrsTitle'] = str(ws['B6'].value or "Reactive MR's Report")
        data['qoeTitle'] = str(ws['B7'].value or 'QOE of PR HFC Network in Green / Optimized')
    else:
        # Defaults
        data['weekLabel'] = 'W19 · 2026'
        data['wkNote'] = 'W5–W19 = 2026'
        data['pnmTitle'] = 'PNM Nodes Completed'
        data['tapsTitle'] = 'SOLI5 Taps Cleaned'
        data['leaksTitle'] = 'Reactive Leaks Reports'
        data['mrsTitle'] = "Reactive MR's Report"
        data['qoeTitle'] = 'QOE of PR HFC Network in Green / Optimized'
    
    return data

def embed_data_in_html(html_template, data):
    """Embed data at END of script so all functions are defined first"""
    # Create the injection script
    data_injection = f"window.INJECTED_DATA = {json.dumps(data)};"
    init_code = "renderAll(window.INJECTED_DATA);"
    
    # Find the LAST </script> tag (not the first) and inject BEFORE it
    last_script_close = html_template.rfind('</script>')
    
    if last_script_close != -1:
        # Insert before the last </script>
        html_template = (
            html_template[:last_script_close] +
            f"\n{data_injection}\n{init_code}\n" +
            html_template[last_script_close:]
        )
    else:
        # Fallback: add new script tag before </body>
        html_template = html_template.replace(
            '</body>',
            f'<script>{data_injection}\n{init_code}</script>\n</body>'
        )
    
    return html_template

def main():
    """Main function"""
    excel_file = Path('dashboard_data.xlsx')
    template_file = Path('liberty_dashboard_template.html')
    output_file = Path('liberty_dashboard.html')
    
    # Check for template
    if not template_file.exists():
        print(f"Error: Template file not found: {template_file}")
        print("Make sure 'liberty_dashboard_template.html' is in the same directory")
        sys.exit(1)
    
    # Check for Excel file
    if not excel_file.exists():
        print(f"Error: Excel file not found: {excel_file}")
        print("Make sure 'dashboard_data.xlsx' is in the same directory")
        sys.exit(1)
    
    print(f"📊 Reading data from: {excel_file}")
    data = read_excel_data(str(excel_file))
    
    print(f"📄 Loading template: {template_file}")
    with open(template_file, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    print("🔄 Embedding data...")
    html_content = embed_data_in_html(html_content, data)
    
    print(f"💾 Writing output: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n✅ Success! Dashboard generated at {timestamp}")
    print(f"📍 Output file: {output_file.absolute()}")

if __name__ == '__main__':
    main()
